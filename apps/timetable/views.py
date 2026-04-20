import json
from datetime import timedelta
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from apps.accounts.utils import permission_required_custom, get_user_permissions
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Spacer, Paragraph, Table, TableStyle

from apps.duty.models import ExamType
from .models import Course, ExamTimetable, Program, TimetableCell, TimetableDateConfig


# ── Helpers ───────────────────────────────────────────────────────────────────

def _grid_rows_for_builder():
    rows = []
    for prog in Program.objects.select_related('department').order_by('department__name', 'name'):
        for sem in range(1, prog.total_semesters + 1):
            rows.append({
                'program_id': prog.id,
                'program_name': prog.name,
                'department_name': prog.department.name,
                'semester': sem,
                'row_key': f'{prog.id}-{sem}',
                'program_type': prog.program_type,
            })
    return rows


def _get_date_time_map(timetable):
    """
    Returns a dict of { 'YYYY-MM-DD': {'time': 'HH:MM', 'duration': {'hours': int, 'minutes': int}} } 
    for all configured date overrides. Falls back to ExamType defaults where no override exists.
    """
    configs = {}
    for cfg in TimetableDateConfig.objects.filter(timetable=timetable):
        time_str = cfg.exam_time.strftime('%H:%M') if cfg.exam_time else None
        duration = {
            'hours': cfg.duration_hours or timetable.exam_type.duration_hours or 0,
            'minutes': cfg.duration_minutes or timetable.exam_type.duration_minutes or 0
        }
        configs[str(cfg.date)] = {'time': time_str, 'duration': duration}
    return configs


def _timetable_dates_from_cells_or_range(timetable):
    dates = sorted({
        row['date']
        for row in TimetableCell.objects.filter(timetable=timetable).values('date').distinct()
    })
    if dates:
        return dates
    all_dates = []
    current = timetable.date_from
    while current <= timetable.date_to:
        if current.weekday() != 6:
            all_dates.append(current)
        current += timedelta(days=1)
    return all_dates


def _timetable_matrix(timetable):
    cells = list(
        TimetableCell.objects.filter(timetable=timetable)
        .select_related('program', 'course')
        .order_by('program__department__name', 'program__name', 'semester', 'date')
    )
    dates = _timetable_dates_from_cells_or_range(timetable)
    row_keys = sorted({(c.program_id, c.semester) for c in cells}, key=lambda x: (x[0], x[1]))
    program_ids = [p for p, _ in row_keys]
    programs = {p.id: p for p in Program.objects.filter(id__in=program_ids).select_related('department')}
    labels = {}
    for pid, sem in row_keys:
        p = programs.get(pid)
        labels[(pid, sem)] = f'{p.department.name} — {p.name} (Sem {sem})' if p else f'Program {pid} (Sem {sem})'

    course_lookup = {}
    for c in cells:
        key = (c.program_id, c.semester, c.date)
        text = f'{c.course.code} — {c.course.name}' if c.course and c.course.code else (c.course.name if c.course else '--')
        course_lookup[key] = text

    # Build date → time map (override or default)
    date_time_map = _get_date_time_map(timetable)
    default_time = (
        timetable.exam_type.default_start_time.strftime('%H:%M')
        if timetable.exam_type.default_start_time else None
    )

    return dates, row_keys, labels, course_lookup, date_time_map, default_time


# ── Views ─────────────────────────────────────────────────────────────────────

@login_required
@permission_required_custom('can_view_timetable')
def timetable_view(request):
    sessions = ExamTimetable.objects.all().order_by('-created_at')
    context = {'sessions': sessions, 'permissions': get_user_permissions(request.user)}
    return render(request, 'timetable/list.html', context)


@login_required
@permission_required_custom('can_edit_timetable')
def timetable_builder(request, id=None):
    timetable = get_object_or_404(ExamTimetable, id=id) if id else None

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            exam_type_id = data.get('exam_type')
            date_from = data.get('date_from')
            date_to = data.get('date_to')
            cells = data.get('cells', [])
            # dict of { "YYYY-MM-DD": "HH:MM" or null }
            date_time_overrides = data.get('date_time_overrides', {})
            # dict of { "YYYY-MM-DD": {"hours": int, "minutes": int} }
            date_duration_overrides = data.get('date_duration_overrides', {})

            if not timetable:
                timetable = ExamTimetable.objects.create(
                    exam_type_id=exam_type_id,
                    date_from=date_from,
                    date_to=date_to,
                    created_by=request.user,
                )
            else:
                timetable.exam_type_id = exam_type_id
                timetable.date_from = date_from
                timetable.date_to = date_to
                timetable.save()

            # Rebuild cells
            TimetableCell.objects.filter(timetable=timetable).delete()
            for c in cells:
                TimetableCell.objects.create(
                    timetable=timetable,
                    program_id=c['program_id'],
                    semester=int(c['semester']),
                    date=c['date'],
                    course_id=c.get('course_id') or None,
                )

            # Rebuild date time configs
            TimetableDateConfig.objects.filter(timetable=timetable).delete()
            
            # Get all dates that have either time or duration overrides
            all_dates = set(date_time_overrides.keys()) | set(date_duration_overrides.keys())
            
            for date_str in all_dates:
                time_str = date_time_overrides.get(date_str)
                duration_override = date_duration_overrides.get(date_str, {})
                
                # Create config if there's any override
                if time_str or duration_override:
                    config_data = {
                        'timetable': timetable,
                        'date': date_str,
                    }
                    
                    if time_str:
                        config_data['exam_time'] = time_str
                    
                    if duration_override:
                        config_data['duration_hours'] = duration_override.get('hours', 0)
                        config_data['duration_minutes'] = duration_override.get('minutes', 0)
                    
                    TimetableDateConfig.objects.create(**config_data)

            return JsonResponse({'success': True, 'id': timetable.id})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    exam_types = ExamType.objects.all().order_by('name')
    grid_rows = _grid_rows_for_builder()
    academic_courses = list(
        Course.objects.select_related('program', 'specialisation')
        .order_by('program_id', 'semester', 'code', 'name')
        .values('id', 'name', 'code', 'program_id', 'semester', 'specialisation_id')
    )

    existing_cells = []
    existing_date_times = {}  # { 'YYYY-MM-DD': 'HH:MM' }
    existing_date_durations = {}  # { 'YYYY-MM-DD': {'hours': int, 'minutes': int} }
    existing_exam_type_default_time = None

    if timetable:
        for row in TimetableCell.objects.filter(timetable=timetable).values(
            'program_id', 'semester', 'date', 'course_id'
        ):
            existing_cells.append({
                'program_id': row['program_id'],
                'semester': row['semester'],
                'date': row['date'].isoformat() if row['date'] else None,
                'course_id': row['course_id'],
            })

        for cfg in TimetableDateConfig.objects.filter(timetable=timetable):
            if cfg.exam_time:
                existing_date_times[str(cfg.date)] = cfg.exam_time.strftime('%H:%M')
            
            # Load duration overrides
            if cfg.duration_hours is not None or cfg.duration_minutes is not None:
                existing_date_durations[str(cfg.date)] = {
                    'hours': cfg.duration_hours or 0,
                    'minutes': cfg.duration_minutes or 0
                }

        if timetable.exam_type.default_start_time:
            existing_exam_type_default_time = timetable.exam_type.default_start_time.strftime('%H:%M')

    # Build exam type → default start time map for JS
    exam_type_times = {}
    for et in exam_types:
        exam_type_times[et.id] = {
            'default_start_time': et.default_start_time.strftime('%H:%M') if et.default_start_time else '',
            'duration_hours': et.duration_hours,
            'duration_minutes': et.duration_minutes,
        }

    context = {
        'timetable': timetable,
        'exam_types': exam_types,
        'grid_rows_json': json.dumps(grid_rows, default=str, ensure_ascii=False),
        'academic_courses_json': json.dumps(academic_courses, default=str, ensure_ascii=False),
        'existing_cells_json': json.dumps(existing_cells, default=str, ensure_ascii=False),
        'existing_date_times_json': json.dumps(existing_date_times, ensure_ascii=False),
        'existing_date_durations_json': json.dumps(existing_date_durations, ensure_ascii=False),
        'exam_type_times_json': json.dumps(exam_type_times, ensure_ascii=False),
        'existing_exam_type_default_time': existing_exam_type_default_time or '',
        'permissions': get_user_permissions(request.user),
    }
    return render(request, 'timetable/builder.html', context)


# ── Export helpers ────────────────────────────────────────────────────────────

@login_required
@permission_required_custom('can_view_timetable')
def timetable_export_excel(request, id):
    timetable = get_object_or_404(ExamTimetable, id=id)
    dates, row_keys, labels, course_lookup, date_time_map, default_time = _timetable_matrix(timetable)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Timetable'

    header_fill = PatternFill(start_color='2D2B47', end_color='2D2B47', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    # Build header row
    header = ['Program']
    for d in dates:
        time_str = date_time_map.get(str(d)) or default_time or ''
        label = d.strftime('%d %b %Y')
        if time_str:
            # Convert 24h to 12h for display
            try:
                from datetime import datetime as dt
                label += '\n' + dt.strptime(time_str, '%H:%M').strftime('%I:%M %p')
            except Exception:
                label += '\n' + time_str
        header.append(label)

    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # Data rows
    for pid, sem in row_keys:
        row = [labels[(pid, sem)]]
        for d in dates:
            row.append(course_lookup.get((pid, sem, d), '--'))
        ws.append(row)

    ws.column_dimensions['A'].width = 45
    for i in range(2, len(dates) + 2):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.row_dimensions[1].height = 38

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}.xlsx"'
    return response


@login_required
@permission_required_custom('can_view_timetable')
def timetable_export_pdf(request, id):
    timetable = get_object_or_404(ExamTimetable, id=id)
    dates, row_keys, labels, course_lookup, date_time_map, default_time = _timetable_matrix(timetable)

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=landscape(A4),
        leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'Timetable: {timetable.exam_type.name}', styles['Heading3']),
        Paragraph(f'Date Range: {timetable.date_from} to {timetable.date_to}', styles['Normal']),
        Spacer(1, 10),
    ]

    def _fmt_time(date_str):
        t = date_time_map.get(date_str) or default_time
        if t:
            try:
                from datetime import datetime as dt
                return dt.strptime(t, '%H:%M').strftime('%I:%M %p')
            except Exception:
                return t
        return ''

    data = [['Program'] + [
        f"{d.strftime('%d %b')}\n{_fmt_time(str(d))}" for d in dates
    ]]
    for pid, sem in row_keys:
        r = [labels[(pid, sem)]]
        for d in dates:
            r.append(course_lookup.get((pid, sem, d), '--'))
        data.append(r)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2D2B47')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F4FF')]),
    ]))
    elements.append(table)
    doc.build(elements)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}.pdf"'
    return response