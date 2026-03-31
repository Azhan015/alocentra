from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db import IntegrityError
from django.db.models import ProtectedError
from django.db.models.functions import Lower
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods, require_POST
from apps.accounts.utils import permission_required_custom, get_user_permissions
from apps.timetable.models import Department, Program, Specialisation, Section, Course
from apps.duty.models import ExamType
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string

def _programs_options_qs():
    return Program.objects.select_related('department').order_by('department__name', 'name')


@login_required
@permission_required_custom('can_edit_timetable')
def settings_index(request):
    context = {
        'permissions': get_user_permissions(request.user),
        'departments': Department.objects.all().order_by('name'),
        'programs': _programs_options_qs(),
        'specialisations': Specialisation.objects.select_related('program', 'program__department').order_by(
            'program__department__name', 'program__name', 'name'
        ),
    }
    return render(request, 'college_settings/index.html', context)


@login_required
@permission_required_custom('can_edit_timetable')
def fragment_departments(request):
    departments = Department.objects.all().order_by('name')
    return HttpResponse(
        render_to_string('college_settings/fragments/department_rows.html', {'departments': departments})
    )


@login_required
@permission_required_custom('can_edit_timetable')
def fragment_programs(request):
    programs = _programs_options_qs()
    return HttpResponse(
        render_to_string('college_settings/fragments/program_rows.html', {'programs': programs})
    )


@login_required
@permission_required_custom('can_edit_timetable')
def fragment_specialisations(request):
    program_id = request.GET.get('program_id')
    specs_qs = Specialisation.objects.select_related('program', 'program__department')
    if program_id:
        specs_qs = specs_qs.filter(program_id=program_id)
    specs = specs_qs.order_by('program__department__name', 'program__name', 'name')
    return HttpResponse(
        render_to_string('college_settings/fragments/specialisation_rows.html', {'specialisations': specs})
    )


@login_required
@permission_required_custom('can_edit_timetable')
def fragment_courses(request):
    program_id = request.GET.get('program_id')
    semester = request.GET.get('semester')
    courses_qs = Course.objects.select_related('program', 'program__department', 'specialisation')
    if program_id:
        courses_qs = courses_qs.filter(program_id=program_id)
    if semester:
        try:
            courses_qs = courses_qs.filter(semester=int(semester))
        except (TypeError, ValueError):
            pass
    courses = courses_qs.order_by('program__department__name', 'program__name', 'semester', 'code', 'name')
    return HttpResponse(
        render_to_string('college_settings/fragments/course_rows.html', {'courses': courses})
    )


@login_required
@permission_required_custom('can_edit_timetable')
def fragment_sections(request):
    sections = Section.objects.select_related('program', 'program__department', 'specialisation').order_by(
        'program__department__name', 'program__name', 'semester', 'name'
    )
    return HttpResponse(
        render_to_string('college_settings/fragments/section_rows.html', {'sections': sections})
    )


@login_required
@permission_required_custom('can_edit_timetable')
def fragment_exam_types(request):
    exam_types = ExamType.objects.all().order_by('name')
    return HttpResponse(
        render_to_string('college_settings/fragments/exam_type_rows.html', {'exam_types': exam_types})
    )


@login_required
@permission_required_custom('can_edit_timetable')
def api_settings_reference(request):
    departments = [{'id': d.id, 'name': d.name} for d in Department.objects.order_by('name')]
    programs = [
        {
            'id': p.id,
            'department_id': p.department_id,
            'label': f'{p.department.name} — {p.name}',
            'total_semesters': p.total_semesters,
            'program_type': p.program_type,
        }
        for p in _programs_options_qs()
    ]
    program_types = (
        Program.objects.exclude(program_type__isnull=True)
        .exclude(program_type='')
        .values_list('program_type', flat=True)
        .distinct()
    )
    specialisations = [
        {'id': s.id, 'program_id': s.program_id, 'name': s.name}
        for s in Specialisation.objects.select_related('program').order_by('program_id', 'name')
    ]
    return JsonResponse(
        {
            'departments': departments,
            'programs': programs,
            'specialisations': specialisations,
            'program_types': list(program_types),
        }
    )


@login_required
@permission_required_custom('can_edit_timetable')
def api_programs_by_department(request):
    dept_id = request.GET.get('department_id')
    qs = _programs_options_qs()
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    data = [{'id': p.id, 'label': f'{p.department.name} — {p.name}', 'total_semesters': p.total_semesters} for p in qs]
    return JsonResponse({'programs': data})


# --- Departments ---


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def department_add(request):
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'message': 'Name is required.'})
    Department.objects.create(name=name, created_by=request.user)
    return JsonResponse({'success': True, 'message': 'Department added.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_http_methods(['GET', 'POST'])
def department_edit(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'name': dept.name})
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'message': 'Name is required.'})
    dept.name = name
    dept.save()
    return JsonResponse({'success': True, 'message': 'Department updated.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def department_delete(request, pk):
    Department.objects.filter(pk=pk).delete()
    return JsonResponse({'success': True, 'message': 'Department removed.'})


# --- Programs ---


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def program_add(request):
    department_id = request.POST.get('department_id')
    name = (request.POST.get('name') or '').strip()
    raw_ts = request.POST.get('total_semesters', '6')
    try:
        total_semesters = max(1, min(int(raw_ts), 12))
    except (TypeError, ValueError):
        total_semesters = 6
    if not department_id or not name:
        return JsonResponse({'success': False, 'message': 'Department and program name are required.'})
    get_object_or_404(Department, pk=department_id)
    program_type = (request.POST.get('program_type') or 'UG').strip() or 'UG'
    Program.objects.create(
        department_id=department_id,
        name=name,
        program_type=program_type,
        total_semesters=total_semesters,
    )
    return JsonResponse({'success': True, 'message': 'Program added.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_http_methods(['GET', 'POST'])
def program_edit(request, pk):
    prog = get_object_or_404(Program, pk=pk)
    if request.method == 'GET':
        return JsonResponse(
            {
                'department_id': prog.department_id,
                'name': prog.name,
                'program_type': prog.program_type,
                'total_semesters': prog.total_semesters,
            }
        )
    department_id = request.POST.get('department_id')
    name = (request.POST.get('name') or '').strip()
    raw_ts = request.POST.get('total_semesters', '6')
    try:
        total_semesters = max(1, min(int(raw_ts), 12))
    except (TypeError, ValueError):
        total_semesters = prog.total_semesters
    if not department_id or not name:
        return JsonResponse({'success': False, 'message': 'Department and program name are required.'})
    get_object_or_404(Department, pk=department_id)
    program_type = (request.POST.get('program_type') or prog.program_type or 'UG').strip() or 'UG'
    prog.department_id = department_id
    prog.name = name
    prog.program_type = program_type
    prog.total_semesters = total_semesters
    prog.save()
    return JsonResponse({'success': True, 'message': 'Program updated.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def program_delete(request, pk):
    Program.objects.filter(pk=pk).delete()
    return JsonResponse({'success': True, 'message': 'Program removed.'})


# --- Specialisations ---


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def specialisation_add(request):
    program_id = request.POST.get('program_id')
    name = (request.POST.get('name') or '').strip()
    if not program_id or not name:
        return JsonResponse({'success': False, 'message': 'Program and name are required.'})
    get_object_or_404(Program, pk=program_id)
    Specialisation.objects.create(program_id=program_id, name=name)
    return JsonResponse({'success': True, 'message': 'Specialisation added.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_http_methods(['GET', 'POST'])
def specialisation_edit(request, pk):
    spec = get_object_or_404(Specialisation, pk=pk)
    if request.method == 'GET':
        return JsonResponse({'program_id': spec.program_id, 'name': spec.name})
    program_id = request.POST.get('program_id')
    name = (request.POST.get('name') or '').strip()
    if not program_id or not name:
        return JsonResponse({'success': False, 'message': 'Program and name are required.'})
    get_object_or_404(Program, pk=program_id)
    spec.program_id = program_id
    spec.name = name
    spec.save()
    return JsonResponse({'success': True, 'message': 'Specialisation updated.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def specialisation_delete(request, pk):
    Specialisation.objects.filter(pk=pk).delete()
    return JsonResponse({'success': True, 'message': 'Specialisation removed.'})


# --- Academic courses ---


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def academic_course_add(request):
    program_id = request.POST.get('program_id')
    raw_sem = request.POST.get('semester', '1')
    code = (request.POST.get('code') or '').strip()
    name = (request.POST.get('name') or '').strip()
    apply_all = request.POST.get('apply_all') == '1'
    spec_ids = request.POST.getlist('specialisation_ids')

    if not program_id or not name:
        return JsonResponse({'success': False, 'message': 'Program and course name are required.'})
    prog = get_object_or_404(Program, pk=program_id)
    try:
        semester = max(1, min(int(raw_sem), int(prog.total_semesters)))
    except (TypeError, ValueError):
        semester = 1

    course_type = (request.POST.get('course_type') or 'core').strip().lower() or 'core'
    if course_type not in dict(Course.COURSE_TYPE_CHOICES):
        course_type = 'core'

    if apply_all:
        Course.objects.create(
            program_id=program_id,
            semester=semester,
            code=code,
            name=name,
            course_type=course_type,
            specialisation=None,
        )
    else:
        if not spec_ids:
            return JsonResponse(
                {'success': False, 'message': 'Select at least one specialisation or enable “all sections”.'}
            )
        for sid in spec_ids:
            get_object_or_404(Specialisation, pk=sid, program_id=program_id)
        for sid in spec_ids:
            Course.objects.create(
                program_id=program_id,
                semester=semester,
                code=code,
                name=name,
                course_type=course_type,
                specialisation_id=sid,
            )
    return JsonResponse({'success': True, 'message': 'Course(s) added.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_http_methods(['GET', 'POST'])
def academic_course_edit(request, pk):
    c = get_object_or_404(Course, pk=pk)
    if request.method == 'GET':
        return JsonResponse(
            {
                'program_id': c.program_id,
                'semester': c.semester,
                'code': c.code,
                'name': c.name,
                'course_type': c.course_type,
                'specialisation_id': c.specialisation_id,
            }
        )
    program_id = request.POST.get('program_id')
    raw_sem = request.POST.get('semester', '1')
    code = (request.POST.get('code') or '').strip()
    name = (request.POST.get('name') or '').strip()
    spec_raw = request.POST.get('specialisation_id')
    if not program_id or not name:
        return JsonResponse({'success': False, 'message': 'Program and course name are required.'})
    prog = get_object_or_404(Program, pk=program_id)
    try:
        semester = max(1, min(int(raw_sem), int(prog.total_semesters)))
    except (TypeError, ValueError):
        semester = 1
    spec_id = int(spec_raw) if spec_raw else None
    if spec_id:
        get_object_or_404(Specialisation, pk=spec_id, program_id=program_id)
    course_type = (request.POST.get('course_type') or c.course_type or 'core').strip().lower() or 'core'
    if course_type not in dict(Course.COURSE_TYPE_CHOICES):
        course_type = 'core'
    c.program_id = program_id
    c.semester = semester
    c.code = code
    c.name = name
    c.course_type = course_type
    c.specialisation_id = spec_id
    c.save()
    return JsonResponse({'success': True, 'message': 'Course updated.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def academic_course_bulk_add(request):
    """
    Bulk add multiple Course rows in one request.
    Expects:
      program_id, semester, specialisation_id (optional), course_type,
      items_json: JSON list of {code, name}
    """
    import json

    program_id = request.POST.get('program_id')
    raw_sem = request.POST.get('semester', '1')
    spec_raw = request.POST.get('specialisation_id')
    course_type = (request.POST.get('course_type') or 'core').strip().lower() or 'core'
    items_json = request.POST.get('items_json') or '[]'

    if not program_id:
        return JsonResponse({'success': False, 'message': 'Program is required.'})
    prog = get_object_or_404(Program, pk=program_id)
    try:
        semester = max(1, min(int(raw_sem), int(prog.total_semesters)))
    except (TypeError, ValueError):
        semester = 1
    spec_id = int(spec_raw) if spec_raw else None
    if spec_id:
        get_object_or_404(Specialisation, pk=spec_id, program_id=program_id)

    if course_type not in dict(Course.COURSE_TYPE_CHOICES):
        course_type = 'core'

    try:
        items = json.loads(items_json)
    except Exception:
        items = []

    created = 0
    failed = 0
    errors = []
    for idx, it in enumerate(items, start=1):
        code = (it.get('code') or '').strip()
        name = (it.get('name') or '').strip()
        if not name:
            failed += 1
            errors.append({'row': idx, 'reason': 'Missing course_name'})
            continue
        Course.objects.create(
            program_id=program_id,
            semester=semester,
            code=code,
            name=name,
            course_type=course_type,
            specialisation_id=spec_id,
        )
        created += 1
    return JsonResponse({'success': True, 'created': created, 'failed': failed, 'errors': errors})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def academic_course_delete(request, pk):
    Course.objects.filter(pk=pk).delete()
    return JsonResponse({'success': True, 'message': 'Course removed.'})


# --- Sections ---


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def section_add(request):
    program_id = request.POST.get('program_id')
    spec_raw = request.POST.get('specialisation_id')
    raw_sem = request.POST.get('semester', '1')
    sec_name = (request.POST.get('name') or '').strip()
    if not program_id or not sec_name:
        return JsonResponse({'success': False, 'message': 'Program and section name are required.'})
    prog = get_object_or_404(Program, pk=program_id)
    try:
        semester = max(1, min(int(raw_sem), int(prog.total_semesters)))
    except (TypeError, ValueError):
        semester = 1
    spec_id = int(spec_raw) if spec_raw else None
    if spec_id:
        get_object_or_404(Specialisation, pk=spec_id, program_id=program_id)
    Section.objects.create(
        program_id=program_id,
        specialisation_id=spec_id,
        semester=semester,
        name=sec_name,
    )
    return JsonResponse({'success': True, 'message': 'Section added.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def section_bulk_add(request):
    program_id = request.POST.get('program_id')
    spec_raw = request.POST.get('specialisation_id')
    raw_sem = request.POST.get('semester', '1')
    names_blob = (request.POST.get('names') or '').strip()
    if not program_id or not names_blob:
        return JsonResponse({'success': False, 'message': 'Program and section letters are required.'})
    prog = get_object_or_404(Program, pk=program_id)
    try:
        semester = max(1, min(int(raw_sem), int(prog.total_semesters)))
    except (TypeError, ValueError):
        semester = 1
    spec_id = int(spec_raw) if spec_raw else None
    if spec_id:
        get_object_or_404(Specialisation, pk=spec_id, program_id=program_id)
    parts = [p.strip() for p in names_blob.replace(';', ',').split(',') if p.strip()]
    if not parts:
        return JsonResponse({'success': False, 'message': 'Enter at least one section label.'})
    for sec_name in parts:
        Section.objects.create(
            program_id=program_id,
            specialisation_id=spec_id,
            semester=semester,
            name=sec_name,
        )
    return JsonResponse({'success': True, 'message': f'{len(parts)} section(s) added.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_http_methods(['GET', 'POST'])
def section_edit(request, pk):
    sec = get_object_or_404(Section, pk=pk)
    if request.method == 'GET':
        return JsonResponse(
            {
                'program_id': sec.program_id,
                'specialisation_id': sec.specialisation_id,
                'semester': sec.semester,
                'name': sec.name,
            }
        )
    program_id = request.POST.get('program_id')
    spec_raw = request.POST.get('specialisation_id')
    raw_sem = request.POST.get('semester', '1')
    sec_name = (request.POST.get('name') or '').strip()
    if not program_id or not sec_name:
        return JsonResponse({'success': False, 'message': 'Program and section name are required.'})
    prog = get_object_or_404(Program, pk=program_id)
    try:
        semester = max(1, min(int(raw_sem), int(prog.total_semesters)))
    except (TypeError, ValueError):
        semester = 1
    spec_id = int(spec_raw) if spec_raw else None
    if spec_id:
        get_object_or_404(Specialisation, pk=spec_id, program_id=program_id)
    sec.program_id = program_id
    sec.specialisation_id = spec_id
    sec.semester = semester
    sec.name = sec_name
    sec.save()
    return JsonResponse({'success': True, 'message': 'Section updated.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def section_delete(request, pk):
    Section.objects.filter(pk=pk).delete()
    return JsonResponse({'success': True, 'message': 'Section removed.'})


# --- Exam types ---


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def exam_type_add(request):
    name = (request.POST.get('name') or '').strip()
    raw_h = request.POST.get('duration_hours', '0')
    raw_m = request.POST.get('duration_minutes', '0')
    exam_program_type = (request.POST.get('exam_program_type') or '').strip()
    try:
        h = max(0, min(int(raw_h), 5))
    except (TypeError, ValueError):
        h = 0
    try:
        m = int(raw_m)
        if m not in (0, 15, 30, 45):
            m = 0
    except (TypeError, ValueError):
        m = 0
    if not name:
        return JsonResponse({'success': False, 'message': 'Name is required.'})
    try:
        ExamType.objects.create(
            name=name,
            duration_hours=h,
            duration_minutes=m,
            exam_program_type=exam_program_type,
        )
        return JsonResponse({'success': True, 'message': 'Exam type added.'})
    except IntegrityError:
        return JsonResponse({'success': False, 'message': 'An exam type with the same name, program type and duration already exists.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_http_methods(['GET', 'POST'])
def exam_type_edit(request, pk):
    et = get_object_or_404(ExamType, pk=pk)
    if request.method == 'GET':
        return JsonResponse(
            {
                'name': et.name,
                'duration_hours': et.duration_hours,
                'duration_minutes': et.duration_minutes,
                'exam_program_type': et.exam_program_type,
            }
        )
    name = (request.POST.get('name') or '').strip()
    raw_h = request.POST.get('duration_hours', '0')
    raw_m = request.POST.get('duration_minutes', '0')
    exam_program_type = (request.POST.get('exam_program_type') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'message': 'Name is required.'})
    try:
        h = max(0, min(int(raw_h), 5))
    except (TypeError, ValueError):
        h = 0
    try:
        m = int(raw_m)
        if m not in (0, 15, 30, 45):
            m = 0
    except (TypeError, ValueError):
        m = 0
    et.name = name
    et.duration_hours = h
    et.duration_minutes = m
    et.exam_program_type = exam_program_type
    try:
        et.save()
        return JsonResponse({'success': True, 'message': 'Exam type updated.'})
    except IntegrityError:
        return JsonResponse({'success': False, 'message': 'An exam type with the same name, program type and duration already exists.'})


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def exam_type_delete(request, pk):
    try:
        ExamType.objects.filter(pk=pk).delete()
        return JsonResponse({'success': True, 'message': 'Exam type removed.'})
    except ProtectedError:
        return JsonResponse(
            {'success': False, 'message': 'Cannot delete this exam type because it is still in use.'}
        )


@login_required
@permission_required_custom('can_edit_timetable')
def course_template_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Courses'
    headers = ['department', 'program', 'semester', 'course_code', 'course_name', 'course_type', 'specialisation']
    ws.append(headers)
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename=\"courses_template.xlsx\"'
    return resp


@login_required
@permission_required_custom('can_edit_timetable')
@require_POST
def course_import_excel(request):
    """
    Upload Excel and create Course rows.
    Columns must be: Program, Semester, Code, Name, Type, Specialisation
    """
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'success': False, 'message': 'No file uploaded.'}, status=400)

    try:
        wb = load_workbook(filename=f, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error reading Excel: {str(e)}'}, status=400)

    # Verify headers
    header_row = [str(c.value).strip().lower() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required = ['department', 'program', 'semester', 'code', 'name', 'type', 'specialisation']
    
    # Check if all required headers are present in the first row
    found_all = True
    for r in required:
        if r not in header_row:
            found_all = False
            break
            
    if not found_all:
        return JsonResponse(
            {
                'success': False,
                'message': f'Invalid columns. Required: {", ".join(required)}',
                'found': header_row,
            },
            status=400,
        )

    # Create mapping of head name to column index
    col_map = {name: i for i, name in enumerate(header_row)}

    imported = 0
    failed = 0
    errors = []

    type_map = {lbl.lower(): key for key, lbl in Course.COURSE_TYPE_CHOICES}

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            dept_name = (row[col_map['department']].value or '').strip() if row[col_map['department']].value is not None else ''
            program_name = (row[col_map['program']].value or '').strip() if row[col_map['program']].value is not None else ''
            semester_raw = row[col_map['semester']].value
            code = (row[col_map['code']].value or '').strip() if row[col_map['code']].value is not None else ''
            name = (row[col_map['name']].value or '').strip() if row[col_map['name']].value is not None else ''
            ctype_raw = (row[col_map['type']].value or '').strip() if row[col_map['type']].value is not None else ''
            spec_name = (row[col_map['specialisation']].value or '').strip() if row[col_map['specialisation']].value is not None else ''

            if not dept_name or not program_name or not name:
                failed += 1
                errors.append({'row': idx, 'reason': 'Missing Department, Program or Name'})
                continue

            dept = Department.objects.filter(name__iexact=dept_name).first()
            if not dept:
                failed += 1
                errors.append({'row': idx, 'reason': f'Department not found: {dept_name}'})
                continue

            prog = Program.objects.filter(department=dept, name__iexact=program_name).first()
            if not prog:
                failed += 1
                errors.append({'row': idx, 'reason': f'Unknown program: {program_name} in department {dept_name}'})
                continue

            try:
                sem = int(semester_raw)
            except Exception:
                sem = 1
            if sem < 1 or sem > int(prog.total_semesters):
                failed += 1
                errors.append({'row': idx, 'reason': f'Invalid semester: {semester_raw}'})
                continue

            # Default to core if type not provided or invalid
            ctype_key = type_map.get(ctype_raw.lower())
            if not ctype_key:
                ctype_key = ctype_raw.lower() if ctype_raw else 'core'
                
            if ctype_key not in dict(Course.COURSE_TYPE_CHOICES):
                ctype_key = 'core'

            spec_id = None
            if spec_name:
                spec = Specialisation.objects.filter(program=prog, name__iexact=spec_name).first()
                if not spec:
                    failed += 1
                    errors.append({'row': idx, 'reason': f'Unknown specialisation: {spec_name}'})
                    continue
                spec_id = spec.id

            Course.objects.create(
                program=prog,
                semester=sem,
                code=code,
                name=name,
                course_type=ctype_key,
                specialisation_id=spec_id,
            )
            imported += 1
        except Exception as e:
            failed += 1
            errors.append({'row': idx, 'reason': str(e)})

    return JsonResponse({'success': True, 'imported': imported, 'failed': failed, 'errors': errors})
